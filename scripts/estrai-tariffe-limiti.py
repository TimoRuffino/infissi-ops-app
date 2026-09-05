#!/usr/bin/env python3
"""
Estrae da «CALCOLO NUOVI LIMITI.xlsx» il seed delle tariffe del computo limiti
(shared/limiti/tariffe-seed.json). Una tantum: il foglio NON va nel repository
(13 MB, dati aziendali). Escono SOLO dati: massimali Allegato A, catalogo
prodotti DEI con i prezzi usati da CHECK2 (fogli «Calcolo Automatici A…F»,
colonna PREZZO UNITARIO — NON il foglio «DEI», che è un listino più vecchio),
accessori con la regola di applicazione, controtelai, prezzi delle opere,
coefficienti e regole di detrazione. Le formule vivono nel motore
(server/computo/motore.ts) e sono documentate in
docs/superpowers/specs/2026-09-03-limiti-analisi-fogli-reali.md.

Uso:
  python3 scripts/estrai-tariffe-limiti.py "<percorso>/CALCOLO NUOVI LIMITI .xlsx"
"""
import json
import re
import sys
import unicodedata
import warnings
from datetime import date

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402
from openpyxl.utils import column_index_from_string, get_column_letter  # noqa: E402
from openpyxl.worksheet.formula import ArrayFormula  # noqa: E402

if len(sys.argv) < 2:
    sys.exit("uso: estrai-tariffe-limiti.py <file.xlsx>")

WF = openpyxl.load_workbook(sys.argv[1], data_only=False, read_only=True)
WV = openpyxl.load_workbook(sys.argv[1], data_only=True, read_only=True)


# ── utilità ────────────────────────────────────────────────────────────────
def griglia(nome, r1, r2, c2=260):
    """Formule e valori di un rettangolo: {coord: formula}, {coord: valore}."""
    F, V = {}, {}
    for rf, rv in zip(
        WF[nome].iter_rows(min_row=r1, max_row=r2, max_col=c2),
        WV[nome].iter_rows(min_row=r1, max_row=r2, max_col=c2),
    ):
        for cf, cv in zip(rf, rv):
            if not hasattr(cf, "coordinate"):
                continue
            x = cf.value
            if isinstance(x, ArrayFormula):
                x = x.text
            if x not in (None, ""):
                F[cf.coordinate] = x
            if cv.value not in (None, ""):
                V[cf.coordinate] = cv.value
    return F, V


def col_di(coord):
    return re.match(r"[A-Z]+", coord).group(0)


def riga_di(coord):
    return int(re.search(r"\d+", coord).group(0))


def num(v, d=2):
    return round(float(v), d) if isinstance(v, (int, float)) else None


def slug(testo):
    s = unicodedata.normalize("NFKD", str(testo)).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9]+", "_", s.lower()).strip("_")
    return s[:60]


# ── Massimali Allegato A: 'Calcolo Automatici' B3:D26 ─────────────────────
_, ca = griglia("Calcolo Automatici", 3, 26, 4)
massimali, visti = [], set()
for r in range(3, 27):
    etichetta = str(ca.get(f"B{r}", "")).strip()
    gruppo = etichetta[:1]
    if gruppo == "D":
        gruppo = "C"  # schermature: stesso massimale degli oscuranti (CHECK1 riga 8)
    zona, prezzo = ca.get(f"C{r}"), num(ca.get(f"D{r}"))
    if gruppo in "ABC" and zona and prezzo and (gruppo, zona) not in visti:
        visti.add((gruppo, zona))
        massimali.append({"gruppo": gruppo, "zona": str(zona), "euroMq": prezzo})
assert len(massimali) == 18, len(massimali)

# ── Catalogo prodotti + accessori dai fogli 'Calcolo Automatici A…F' ───────
# Ogni foglio ha 50 blocchi identici (uno per riga di SERRAMENTI); il primo
# blocco (righe 1..N) basta. In riga 4 i marcatori 'Total' aprono i
# sotto-blocchi: [Total][Quantità][Prodotti]…[CODICE][MQ][PREZZO UNITARIO][accessori…].
# La regola di applicazione degli accessori viene dalla formula del totale
# (vedi analisi §3): è dichiarata qui per sotto-blocco.
SOTTOBLOCCHI = {
    # (foglio, colonna Total): (gruppo, famiglia | None=dal nome, regola pct, regola cad)
    ("Calcolo Automatici A", "B"): ("serramento", None, "pct_mq", "cad_pezzo"),
    ("Calcolo Automatici A", "AZ"): ("cassonetto", None, "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici A", "BN"): ("porta_blindata", "blindata", "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici A", "BW"): ("portoncino", "pvc", "pct_mq", "cad_fisso"),
    ("Calcolo Automatici B", "BN"): ("avvolgibile", None, "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici B", "CF"): ("persiana", None, "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici B", "EA"): ("scuro", "legno", "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici C", "U"): ("persiana", None, "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici D", "C"): ("schermatura", "pergola", "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici D", "P"): ("schermatura", "tenda", "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici D", "AD"): ("schermatura", "veneziana", "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici D", "AR"): ("schermatura", "zanzariera", "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici E", "C"): ("serramento", None, "pct_pezzo", "cad_pezzo"),
    ("Calcolo Automatici E", "BS"): ("cassonetto", "legno", "pct_pezzo", "cad_pezzo"),
}
RIGHE_BLOCCO = {"Calcolo Automatici A": 65, "Calcolo Automatici B": 65, "Calcolo Automatici C": 28,
                "Calcolo Automatici D": 100, "Calcolo Automatici E": 78}


def famiglia_serramento(nome):
    n = nome.upper()
    if n.startswith("LEGNO - ALLUMINIO"):
        return "legno_alluminio"
    if n.startswith("ALLUMINIO - LEGNO"):
        return "alluminio_legno"
    if n.startswith("LEGNO"):
        return "legno"
    if n.startswith("ALLUMINIO"):
        return "alluminio"
    if n.startswith("PVC"):
        return "pvc"
    if n.startswith("VELUX"):
        return "velux"
    return "altro"


def famiglia_generica(gruppo, nome):
    n = nome.lower()
    if gruppo == "cassonetto":
        if "monoblocco" in n:
            return "monoblocco"
        if "coibentazione" in n:
            return "coibentazione"
        if "oltre" in n:
            return "pvc_oltre_110"
        return "pvc_fino_110"
    if gruppo in ("persiana", "avvolgibile", "scuro"):
        if "alluminio pvc" in n:
            return "alluminio_pvc"
        for f in ("pvc", "legno", "alluminio", "acciaio", "velux"):
            if f in n:
                return f
        return "altro"
    return "altro"


def zone_da_nome(nome):
    m = re.search(r"zon[ae] climatich?e?\s*([A-F](?:\s*-\s*[A-F])?)", nome, re.I)
    if not m:
        return None
    z = m.group(1).replace(" ", "").upper()
    return z.split("-") if "-" in z else [z]


def n_ante(nome):
    n = nome.lower()
    if "2 ante" in n or "due ante" in n:
        return 2
    if "3 ante" in n:
        return 3
    if "4 ante" in n:
        return 4
    return 1


def classe_mq(nome):
    """Cassonetti: intervallo di mq per pezzo che seleziona la voce (CHECK2 AZ)."""
    n = nome.lower()
    m = re.search(r"(\d{3})\s*×\s*40", n)
    if m:
        return {"100": (0, 0.51), "150": (0.51, 0.71), "200": (0.71, 0.91), "250": (0.91, 1.11), "300": (1.11, None)}[m.group(1)]
    if "fino a 1 mq" in n:
        return (0, 1.0)
    m = re.search(r"da (\d[,.]\d+) a (\d[,.]\d+) mq", n)
    if m:
        return (float(m.group(1).replace(",", ".")), float(m.group(2).replace(",", ".")))
    return None


def regola_accessorio(formula, valore, regola_pct, regola_cad, colonna_prezzo):
    """(regola, valore, moltiplicatore) dalla cella accessorio della riga prodotto.

    `=J5*0.15` / `=0.02*J16` → percentuale del prezzo unitario; `=M5*2` → costante
    per anta (valore base, moltiplicatore 2); `=120`, `=M5`, costante → importo fisso.
    """
    f = str(formula).replace(" ", "") if formula is not None else ""
    if f.startswith("="):
        # percentuale del prezzo unitario: `=J5*0.15`, `=0.02*J16`, `=CL82*1` (100 %), `=1*H5`
        m_pct = re.fullmatch(r"=(?:(\d*\.?\d+)\*)?([A-Z]+\d+)(?:\*(\d*\.?\d+))?", f)
        if m_pct and (m_pct.group(1) or m_pct.group(3)) and col_di(m_pct.group(2)) == colonna_prezzo:
            return regola_pct, round(float(m_pct.group(1) or m_pct.group(3)) * 100, 2), 1
        m_molt = re.fullmatch(r"=([A-Z]+\d+)\*(\d+)", f)
        if m_molt and isinstance(valore, (int, float)):
            return regola_cad, num(valore / int(m_molt.group(2))), int(m_molt.group(2))
        if re.fullmatch(r"=[A-Z]+\d+", f) or re.fullmatch(r"=\d+(\.\d+)?", f):
            return regola_cad, num(valore), 1
        return None
    if isinstance(formula, (int, float)):
        return regola_cad, num(formula), 1
    return None


prodotti, accessori = [], []
codici_prodotto, codici_accessorio = set(), set()

for (foglio, col_total), (gruppo, famiglia_fissa, regola_pct, regola_cad) in SOTTOBLOCCHI.items():
    F, V = griglia(foglio, 1, RIGHE_BLOCCO[foglio])
    c0 = column_index_from_string(col_total)
    totali = sorted(column_index_from_string(col_di(k)) for k, v in V.items() if riga_di(k) == 4 and v == "Total")
    fine = next((t - 1 for t in totali if t > c0), 260)
    etich = {column_index_from_string(col_di(k)): v for k, v in V.items() if riga_di(k) == 4 and c0 <= column_index_from_string(col_di(k)) <= fine}

    def trova(label):
        return next((ci for ci, v in etich.items() if isinstance(v, str) and v.strip().upper() == label), None)

    c_prod, c_cod, c_prezzo = trova("PRODOTTI"), trova("CODICE"), trova("PREZZO UNITARIO")
    assert c_prod and c_prezzo, (foglio, col_total)
    L_prod, L_cod, L_prezzo = get_column_letter(c_prod), get_column_letter(c_cod) if c_cod else None, get_column_letter(c_prezzo)
    unita = "m" if str(etich.get(c_prezzo - 1, "")).strip() == "M" else ("cad" if gruppo in ("cassonetto", "porta_blindata") else "mq")

    # colonne accessorio: nome in riga 3 tra prezzo+1 e fine; 'Codice' e 'Perimetro' sono colonne di servizio
    colonne_acc = []
    for ci in range(c_prezzo + 1, fine + 1):
        L = get_column_letter(ci)
        nome_acc = V.get(f"{L}3")
        if not isinstance(nome_acc, str) or nome_acc.strip() in ("", "Codice", "Perimetro", "Altezza", "Larghezza", "Altezza (H)", "Larghezza (L)", "Intervalos"):
            continue
        succ = str(V.get(f"{get_column_letter(ci + 1)}3", "")).strip()
        perimetro = succ == "Perimetro"
        L_codice = get_column_letter(ci + (2 if perimetro else 1))
        if str(V.get(f"{L_codice}3", "")).strip() != "Codice":
            L_codice = None
        colonne_acc.append((L, nome_acc.strip(), L_codice, perimetro))

    prima_riga_valida = None
    for r in range(5, RIGHE_BLOCCO[foglio]):
        nome = V.get(f"{L_prod}{r}")
        prezzo = num(V.get(f"{L_prezzo}{r}"))
        if not isinstance(nome, str) or not nome.strip() or prezzo is None:
            continue
        nome = " ".join(nome.split())
        codice = str(V.get(f"{L_cod}{r}") or "").strip() if L_cod else ""
        if not codice:
            codice = f"{gruppo}:{slug(nome)}"
        if codice in codici_prodotto:
            continue  # stesso prodotto già preso da un altro foglio (B ripete A, F ripete E…)
        codici_prodotto.add(codice)
        famiglia = famiglia_fissa or (famiglia_serramento(nome) if gruppo == "serramento" else famiglia_generica(gruppo, nome))
        p = {
            "codice": codice, "gruppo": gruppo, "famiglia": famiglia, "nome": nome, "prezzo": prezzo, "unita": unita,
            "foglio": foglio.replace("Calcolo Automatici ", ""),
        }
        if gruppo == "serramento":
            p["zone"] = zone_da_nome(nome)
            p["nAnte"] = n_ante(nome)
            p["portafinestra"] = "portafinestra" in nome.lower() or "portefinestre" in nome.lower()
            # Minimo di fatturazione 1 mq: nel foglio è applicato SOLO ai blocchi A/B (PVC/alluminio), non al legno (E).
            p["minimoMq"] = 1 if p["foglio"] in ("A", "B") else None
        if gruppo == "avvolgibile":
            p["minimoMq"] = 1.8
        if gruppo == "cassonetto" and unita == "cad":
            cl = classe_mq(nome)
            if cl:
                p["mqPezzoMin"], p["mqPezzoMax"] = cl
        if gruppo == "schermatura":
            def intervallo(L):
                v = str(V.get(f"{L}{r}", "")).strip()
                return v if v and any(ch in v for ch in "<>[(") else None
            p["intervalloL"] = intervallo(get_column_letter(c_prod + 4))
            p["intervalloH"] = intervallo(get_column_letter(c_prod + 5))
        prodotti.append(p)
        prima_riga_valida = prima_riga_valida or r

    # accessori: un'entrata per colonna, valore preso dalla prima riga prodotto con contenuto (le percentuali sono costanti per colonna)
    intest4 = {column_index_from_string(col_di(k)): v for k, v in V.items() if riga_di(k) == 4}
    TITOLI = {"SERRAMENTI PVC": ["pvc", "velux"], "SERRAMENTI ALLUMINIO": ["alluminio"], "PERSIANE PVC": ["pvc"],
              "PERSIANE LEGNO": ["legno"], "PERSIANE ALLUMINIO": ["alluminio"], "SCURI LEGNO": ["legno"],
              "LEGNO": ["legno"], "LEGNO - ALLUMINIO": ["legno_alluminio", "alluminio_legno"],
              "MOTORE": None, "SOVRAPREZZO AVVOLGIBILI": None, "CASSONETTI": None, "PORTONCINI": None}
    for L, nome_acc, L_codice, perimetro in colonne_acc:
        r = None
        for rr in range(prima_riga_valida or 5, RIGHE_BLOCCO[foglio]):
            fx0, vx0 = F.get(f"{L}{rr}"), V.get(f"{L}{rr}")
            if (isinstance(fx0, str) and fx0.startswith("=")) or (isinstance(vx0, (int, float)) and vx0 != 0):
                r = rr
                break
        if r is None:
            continue
        fx, vx = F.get(f"{L}{r}"), V.get(f"{L}{r}")
        if perimetro:
            regola = ("m_perimetro", num(vx), 1)
        else:
            regola = regola_accessorio(fx, vx, regola_pct, regola_cad, L_prezzo)
        if regola is None or regola[1] in (None, 0):
            continue
        codice_dei = str(V.get(f"{L_codice}{r}") or "").strip() if L_codice else ""
        nome_norm = " ".join(nome_acc.split())
        if nome_norm.lower().startswith("incollaggio"):
            regola = ("cad_anta", regola[1], 1)
        solo_pf = nome_norm.lower().startswith("soglia ribassata per portefinestre")
        # famiglia: l'ULTIMO titolo di riga 4 a sinistra della colonna (K4 'Serramenti PVC', AA4 'Serramenti Alluminio', …)
        ci = column_index_from_string(L)
        titolo = next((str(v).upper() for c, v in sorted(intest4.items(), reverse=True)
                       if c <= ci and isinstance(v, str) and str(v).upper() in TITOLI), None)
        fam = TITOLI.get(titolo) if titolo else None
        if fam is None:
            fam = sorted({p["famiglia"] for p in prodotti if p["gruppo"] == gruppo})
        # Moltiplicatori che vivono nella formula del totale, non nella cella (CHECK2 CF: cardini ×2).
        molt = regola[2]
        if gruppo == "persiana" and codice_dei == "C25084-c":
            molt = 2
        chiave_dedup = (gruppo, codice_dei or nome_norm.lower(), tuple(fam))
        if chiave_dedup in codici_accessorio:
            continue
        codici_accessorio.add(chiave_dedup)
        base = codice_dei or slug(nome_norm)
        codice = f"{gruppo}.{base}" if not any(a["codice"] == f"{gruppo}.{base}" for a in accessori) else f"{gruppo}.{base}.{fam[0]}"
        accessori.append({
            "codice": codice, "codiceDei": codice_dei or None, "nome": nome_norm, "gruppo": gruppo,
            "famiglie": fam, "regola": regola[0], "valore": regola[1], "moltiplicatore": molt,
            "soloPortafinestra": solo_pf, "foglio": foglio.replace("Calcolo Automatici ", ""),
        })

assert len(prodotti) > 150, len(prodotti)
assert len(accessori) > 40, len(accessori)

# ── Controtelai: 'Calcolo Automatici' L344:N374 (famiglia = riga senza prezzo)
_, ct = griglia("Calcolo Automatici", 344, 374, 14)
controtelai, famiglia_ct, unita_ct = [], None, None
# Il foglio scrive «acciao» (refuso) accanto ad «acciaio»: si riconosce la radice «accia».
UNITA_CT = {"accia": "mq", "alluminio": "cad", "legno": "m"}
for r in range(344, 375):
    l, m, n = ct.get(f"L{r}"), ct.get(f"M{r}"), num(ct.get(f"N{r}"))
    if l and n is None:
        famiglia_ct = " ".join(str(l).split())
        chiave = next((k for k in UNITA_CT if k in famiglia_ct.lower()), "legno")
        unita_ct = UNITA_CT[chiave]
    elif l and n is not None:
        controtelai.append({"codice": str(l).replace(" ", ""), "famiglia": famiglia_ct, "variante": " ".join(str(m).split()),
                            "unita": unita_ct, "prezzo": n, "minimoMq": 1.2 if unita_ct == "mq" else None})
assert len(controtelai) == 22, len(controtelai)

# ── Opere: CHECK1 E22:E43 con codice DEI in N ──────────────────────────────
_, ck = griglia("CHECK1", 22, 43, 14)
OPERE = [
    (22, "rilievo_pezzo", "h"), (23, "rilievo_foro", "h"), (24, "progettazione", "h"),
    (25, "sviluppo_ordine", "h"), (26, "protezione", "h"), (27, "rimozione_serramenti", "mq"),
    (28, "rimozione_tapparelle", "mq"), (29, "smaltimento", "fisso"), (30, "trasporto", "km"),
    (31, "tiro_piano", "h"), (32, "assistenza_muraria", "m"), (33, "posa", "h"),
    (34, "pulizia", "h"), (35, "spese_professionali", "fisso"),
    (39, "altri_servizi", "%"), (40, "assistenze_murarie_eventuali", "h"),
    (41, "dime", "mq"), (42, "piattaforma", "giornata"), (43, "permessi_suolo", "giornata"),
]
opere = []
for r, codice, unita in OPERE:
    prezzo = num(ck.get(f"E{r}"))
    if r == 42:
        prezzo = round(64.74 * 8, 2)  # E42 = 64,74 × 8
    opere.append({
        "codice": codice, "gruppo": "eventuali" if r >= 39 else "opere",
        "descrizione": " ".join(str(ck.get(f"B{r}") or "").split()),
        "codiceDei": str(ck.get(f"N{r}") or "").split("\n")[0].strip() or None,
        "unita": unita, "prezzo": prezzo,
        # CHECK2 (T46) esclude le opere già comprese nel prezzo DEI «opere compiute».
        "esclusaDaCheck2": codice in ("sviluppo_ordine", "trasporto", "posa"),
        # Rilievo al pezzo è l'alternativa al rilievo a foro; spese professionali ed eventuali entrano solo se richiesti.
        "inclusaDefault": codice not in ("rilievo_pezzo", "spese_professionali") and r < 39,
    })

# ── Coefficienti trascritti dalle formule (CHECK1 H22:H43, Tempi) ──────────
coefficienti = {
    "oreTiro": {"serramento": 0.5, "cassonetto": 0.25, "tapparella": 0.25, "persiana": 0.25, "scuro": 0.25,
                "porta_blindata": 0.5, "portoncino": 0.5, "schermatura": 0.25, "zanzariera": 0.25, "tenda": 1, "pergola": 2,
                "materialiPosa": 1 / 3},
    "orePosa": {"serramento": 3, "cassonetto": 1, "oscurante": 1.5, "schermatura": 1.5, "zanzariera": 1.5, "tenda": 4,
                "pergola": 16, "porta_blindata": 3, "portoncino": 3},
    "oreGiornata": 8, "euroKm": 0.7, "installatori": 2, "maggiorazionePianoOltre": 4, "maggiorazionePiano": 1.3,
    "puliziaFissoEuro": 50,
    # H29: 150 + 104,69 × 0,1 × mq serramenti + 0,35 × mq cassonetti + 0,05 × mq oscuranti
    #      + 100 × 0,025 × mq serramenti + 0,015 × mq cassonetti + 0,0125 × mq oscuranti
    # (la precedenza del foglio moltiplica 104,69 e 100 SOLO per il primo termine: si riproduce così com'è).
    "smaltimentoBaseEuro": 150, "smaltimentoEuroMc": 104.69, "smaltimentoEuroOnere": 100,
    "smaltimentoMcSerramento": 0.1, "smaltimentoMcCassonetto": 0.35, "smaltimentoMcOscurante": 0.05,
    "smaltimentoOnereSerramento": 0.025, "smaltimentoOnereCassonetto": 0.015, "smaltimentoOnereOscurante": 0.0125,
    "speseProfessionaliPct": 0.04, "speseProfessionaliMinEuro": 600, "altriServiziPct": 0.02,
    "controtelaiMinMq": 1.2,
    # IVA agevolata (10 %) usata per stimare l'imponibile da un pattuito lordo prima della fattura.
    "ivaAgevolata": 0.10,
    # Avvolgibili (CHECK2 BS): mq = mq serramento + 0,05 × (L + 0,25) + 0,25 × (H + 0,05), minimo 1,8 mq
    "avvolgibileExtraL": 0.05, "avvolgibileExtraLOffset": 0.25, "avvolgibileExtraH": 0.25, "avvolgibileExtraHOffset": 0.05,
}
# Aliquote per anno di firma. `percentualeDetrazione` prende la riga con
# l'anno più alto <= anno della firma: senza il 2025 un contratto firmato in
# quell'anno non aveva alcuna percentuale e il detraibile restava «—».
# 2025 e 2026: 50 % prima casa, 36 % altri immobili. 2027: 36 % e 30 %
# (legge di bilancio 2025, art. 1 c. 54-55).
detrazioni = [
    {"tipo": "ristrutturazione", "immobile": "prima_casa", "anno": 2025, "pct": 50},
    {"tipo": "ristrutturazione", "immobile": "altro", "anno": 2025, "pct": 36},
    {"tipo": "ecobonus", "immobile": "prima_casa", "anno": 2025, "pct": 50},
    {"tipo": "ecobonus", "immobile": "altro", "anno": 2025, "pct": 36},
    {"tipo": "ristrutturazione", "immobile": "prima_casa", "anno": 2026, "pct": 50},
    {"tipo": "ristrutturazione", "immobile": "altro", "anno": 2026, "pct": 36},
    {"tipo": "ecobonus", "immobile": "prima_casa", "anno": 2026, "pct": 50},
    {"tipo": "ecobonus", "immobile": "altro", "anno": 2026, "pct": 36},
    {"tipo": "ristrutturazione", "immobile": "prima_casa", "anno": 2027, "pct": 36},
    {"tipo": "ristrutturazione", "immobile": "altro", "anno": 2027, "pct": 30},
    {"tipo": "ecobonus", "immobile": "prima_casa", "anno": 2027, "pct": 36},
    {"tipo": "ecobonus", "immobile": "altro", "anno": 2027, "pct": 30},
]
# «accessorio» NON è un bene significativo (P3-R7): coprifili, maniglie e
# simili viaggiano come «altri beni» nelle fatture reali, e contarli fra i
# significativi ridurrebbe l'IVA agevolata sul resto della fornitura.
bene_default = {c: True for c in ["serramento_pvc", "serramento_alluminio", "serramento_legno", "serramento_legno_alluminio",
                                  "cassonetto", "tapparella", "persiana", "scuro", "schermatura", "zanzariera", "tenda",
                                  "pergola", "porta_blindata", "portoncino", "porta_interna"]}
bene_default.update({"accessorio": False, "controtelaio": False, "altro": False})

seed = {
    "versione": date.today().isoformat(), "fonte": "CALCOLO NUOVI LIMITI.xlsx", "validoDal": "2022-04-15",
    "massimali": massimali, "prodotti": prodotti, "accessori": accessori, "controtelai": controtelai, "opere": opere,
    "coefficienti": coefficienti, "detrazioni": detrazioni, "beneSignificativoDefault": bene_default,
}
with open("shared/limiti/tariffe-seed.json", "w", encoding="utf-8") as f:
    json.dump(seed, f, ensure_ascii=False, indent=1)
print({k: (len(v) if isinstance(v, list) else "ok") for k, v in seed.items()})
from collections import Counter  # noqa: E402
print("prodotti per gruppo:", Counter(p["gruppo"] for p in prodotti))
print("accessori per gruppo:", Counter(a["gruppo"] for a in accessori))
