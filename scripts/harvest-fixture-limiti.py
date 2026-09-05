#!/usr/bin/env python3
"""
Trasforma una copia compilata del foglio «CALCOLO NUOVI LIMITI» (.xlsm) in un
caso d'oro per i test del motore limiti (server/computo/__fixtures__/
casi-reali.json). I fogli NON entrano nel repository (13 MB, dati di clienti):
esce SOLO il caso anonimo, cioè misure, codici DEI del seed, prezzi di riga e
totali del foglio. Nessun nome, indirizzo o nome di file compare nell'output:
il caso si chiama come dice `--nome`.

Cosa legge (celle verificate su 19 fogli reali del 2026):
  INIZIO   H11 zona climatica, E17 distanza km, E19 piano.
  SERRAMENTI  righe 7..56 dei blocchi A (col. C..AG), B (AH..CQ), C (CR..EE)
              e D (EF..EO): quantità, L, H, mq, gruppo, prodotto, accessori
              «Sì» (nomi nelle intestazioni di riga 5) e prezzo di riga.
  Stampa   un blocco di 75 righe per riga di SERRAMENTI: D «Limite Costo» (DEI
           della riga) e H il CODICE DEI già risolto dal foglio (per i
           cassonetti è la classe scelta, per le schermature l'intervallo).
  CHECK1   H6/H7/H8 massimali, H11..H18 controtelai, H22..H35 opere,
           H39..H43 eventuali, T6 totale DEI, G46/H46/T46 fattura e check.

Cosa NON legge: B7/B9/B11 di INIZIO (nominativo, indirizzo, comune) e ogni
altra cella con dati personali.

Regole di traduzione foglio → caso (le divergenze note dal motore sono nel
report `.superpowers/sdd/2026-09-05-harvest-fixture-limiti/report.md`):
  - `attesi.voci` porta i limiti del foglio voce per voce (massimali, opere,
    eventuali e `dei_riga_n` dalla colonna D di «Stampa»): sono il vero
    contenuto del caso, indipendenti da cosa è stato fatturato.
  - `attesi.check1`/`check2` sommano gli stessi numeri del foglio con la
    regola di inclusione del MOTORE (opere ordinarie + rilievo scelto +
    spese professionali/eventuali se richiesti). `attesi.foglio` conserva
    H46/T46 così come li calcola il foglio e l'elenco delle opere su cui le
    due inclusioni divergono: il foglio somma solo ciò che è stato davvero
    fatturato (colonna G «Da fattura»), scelta che `OpzioniComputo` oggi non
    sa rappresentare.
  - `parametri.pattuitoCent` = G46 − G35 (il totale fatturato senza le spese
    professionali): è la stessa base con cui il foglio calcola H35.

Uso:
  python3 scripts/harvest-fixture-limiti.py <file.xlsm> --nome fattura-62-2026 \
      --detrazione ristrutturazione [--pct 50] [--salta "motivo"] [--zitto]

Stampa il caso JSON su stdout e la diagnostica (nomi non mappati, codici in
disaccordo con «Stampa», situazioni che il motore non modella) su stderr.
"""
import argparse
import json
import re
import sys
import unicodedata
import warnings
from difflib import SequenceMatcher
from pathlib import Path

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402
from openpyxl.utils import column_index_from_string, get_column_letter  # noqa: E402

RADICE = Path(__file__).resolve().parents[1]
SEED = json.loads((RADICE / "shared/limiti/tariffe-seed.json").read_text(encoding="utf-8"))

# ── Struttura del foglio SERRAMENTI ────────────────────────────────────────
# Un blocco per gruppo dell'Allegato A, affiancati sulle stesse righe 7..56.
# Gli oscuranti dei blocchi B/C sono terne (colonna del prodotto, prima e
# ultima colonna dei suoi accessori).
BLOCCHI = {
    "A": dict(quantita="C", l="D", h="F", mq="H", gruppo="I", tipologia="J", prezzo="AF",
              accessori=("K", "AE"), oscuranti={}),
    "B": dict(quantita="AH", l="AI", h="AK", mq="AM", gruppo="AN", tipologia="AO", prezzo="CP",
              accessori=("AP", "BJ"),
              oscuranti={"tapparella": ("BK", "BL", "BO"), "persiana": ("BP", "BQ", "CJ"),
                         "scuro": ("CK", "CL", "CO")}),
    "C": dict(quantita="CR", l="CS", h="CU", mq="CW", gruppo="CX", tipologia=None, prezzo="ED",
              accessori=None,
              oscuranti={"tapparella": ("CY", "CZ", "DC"), "persiana": ("DD", "DE", "DX"),
                         "scuro": ("DY", "DZ", "EC")}),
    "D": dict(quantita="EF", l="EG", h="EI", mq="EK", gruppo="EL", tipologia="EM", prezzo="EN",
              accessori=None, oscuranti={}),
}
# Blocchi E ed F (legno e legno + oscuranti): nessuno dei fogli reali del 2026
# li usa, quindi la posizione della colonna del prodotto non è verificata.
# Meglio fermarsi che produrre un caso d'oro sbagliato.
QUANTITA_LEGNO = {"E": "EP", "F": "GE"}

# Gruppo dichiarato nella colonna «Prodotto» → oscurante abbinato (blocchi B/F).
OSCURANTE_DEL_GRUPPO = {"serrtapp": "tapparella", "serrpers": "persiana",
                        "serrscuri": "scuro", "portoncinopers": "persiana"}
# Blocco C: la colonna «Tipologia» dice quale famiglia di oscurante è la riga.
OSCURANTE_SOLO = {"tapparelle": "tapparella", "persiane": "persiana", "scuri": "scuro"}

# Opere di CHECK1 nell'ordine delle righe 22..35 e 39..43.
OPERE_ORDINARIE = ["rilievo_pezzo", "rilievo_foro", "progettazione", "sviluppo_ordine",
                   "protezione", "rimozione_serramenti", "rimozione_tapparelle", "smaltimento",
                   "trasporto", "tiro_piano", "assistenza_muraria", "posa", "pulizia",
                   "spese_professionali"]
OPERE_EVENTUALI = ["altri_servizi", "assistenze_murarie_eventuali", "dime", "piattaforma",
                   "permessi_suolo"]
RIGA_OPERA = {c: 22 + i for i, c in enumerate(OPERE_ORDINARIE)}
RIGA_OPERA.update({c: 39 + i for i, c in enumerate(OPERE_EVENTUALI)})
RIGHE_CONTROTELAIO = [11, 13, 14, 17, 18]

CODICE_DEI = re.compile(r"^[A-Z]\d{4,5}(-[a-z])?$")


# ── utilità ────────────────────────────────────────────────────────────────
def num(v):
    """Il numero della cella, o None: i fogli salvano anche i numeri come testo."""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip().replace(",", "."))
        except ValueError:
            return None
    return None


def testo(v):
    return "" if v is None else str(v).strip()


def chiave(s):
    """Nome confrontabile: senza accenti, μ→u, solo lettere e cifre minuscole."""
    s = unicodedata.normalize("NFKD", str(s)).replace("μ", "u").replace("µ", "u")
    s = s.encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def simile(a, b):
    return SequenceMatcher(None, a, b).ratio()


def cella(righe, r, colonna):
    i = column_index_from_string(colonna) - 1
    if r - 1 >= len(righe) or i >= len(righe[r - 1]):
        return None
    return righe[r - 1][i]


def sì(v):
    return chiave(v) in ("si", "s")


class Diagnostica:
    def __init__(self, zitto):
        self.zitto = zitto
        self.nomi_non_mappati = []
        self.avvisi = []

    def avvisa(self, messaggio):
        self.avvisi.append(messaggio)
        if not self.zitto:
            print(f"  ! {messaggio}", file=sys.stderr)

    def nome_non_mappato(self, cosa, nome, dettaglio=""):
        self.nomi_non_mappati.append({"cosa": cosa, "nome": nome, "dettaglio": dettaglio})
        self.avvisa(f"{cosa} «{nome}» non mappato sul seed{(' — ' + dettaglio) if dettaglio else ''}")


# ── catalogo: nome del foglio → voce del seed ─────────────────────────────
def scegli_per_nome(nome, candidati, dia, cosa):
    """Voce del seed con quel nome: uguale, prefisso, oppure la più simile."""
    if not candidati:
        return None
    k = chiave(nome)
    esatti = [c for c in candidati if chiave(c["nome"]) == k]
    if len(esatti) == 1:
        return esatti[0]
    if esatti:
        dia.avvisa(f"{cosa} «{nome}»: {len(esatti)} voci con lo stesso nome, presa {esatti[0]['codice']}")
        return esatti[0]
    prefisso = [c for c in candidati if chiave(c["nome"]).startswith(k)]
    if len(prefisso) == 1:
        return prefisso[0]
    if prefisso:
        dia.avvisa(f"{cosa} «{nome}»: {len(prefisso)} voci con lo stesso prefisso "
                   f"({', '.join(c['codice'] for c in prefisso[:6])}) — serve il codice del foglio")
        return None
    migliore = max(candidati, key=lambda c: simile(k, chiave(c["nome"])))
    punteggio = simile(k, chiave(migliore["nome"]))
    if punteggio >= 0.9:
        dia.avvisa(f"{cosa} «{nome}» → {migliore['codice']} «{migliore['nome']}» (somiglianza {punteggio:.2f})")
        return migliore
    return None


def prodotto_per_codice(codice):
    return next((p for p in SEED["prodotti"] if p["codice"] == codice), None)


CATEGORIA_DI = {
    ("serramento", "pvc"): "serramento_pvc",
    ("serramento", "alluminio"): "serramento_alluminio",
    ("serramento", "velux"): "serramento_pvc",
    ("serramento", "legno"): "serramento_legno",
    ("serramento", "legno_alluminio"): "serramento_legno_alluminio",
    ("serramento", "alluminio_legno"): "serramento_legno_alluminio",
    ("cassonetto", None): "cassonetto",
    ("avvolgibile", None): "tapparella",
    ("persiana", None): "persiana",
    ("scuro", None): "scuro",
    ("porta_blindata", None): "porta_blindata",
    ("portoncino", None): "portoncino",
    ("schermatura", "tenda"): "tenda",
    ("schermatura", "pergola"): "pergola",
    ("schermatura", "zanzariera"): "zanzariera",
    ("schermatura", "veneziana"): "schermatura",
}


def categoria_di(prodotto):
    """Categoria della riga di contratto per il prodotto DEI scelto."""
    g, f = prodotto["gruppo"], prodotto.get("famiglia")
    return CATEGORIA_DI.get((g, f)) or CATEGORIA_DI.get((g, None))


def accessori_marcati(righe_serr, intestazioni, r, prima, ultima, gruppo, dia, etichetta):
    """Codici del seed degli accessori con «Sì» fra due colonne, per un gruppo."""
    candidati = [a for a in SEED["accessori"] if a["gruppo"] == gruppo]
    codici = []
    for c in range(column_index_from_string(prima), column_index_from_string(ultima) + 1):
        colonna = get_column_letter(c)
        if not sì(cella(righe_serr, r, colonna)):
            continue
        nome = testo(intestazioni[c - 1]) if c - 1 < len(intestazioni) else ""
        voce = scegli_per_nome(nome, candidati, dia, f"accessorio ({etichetta})")
        if not voce:
            dia.nome_non_mappato("accessorio", nome, f"riga {r}, gruppo {gruppo}")
            continue
        codici.append(voce["codice"])
    return codici


# ── lettura del foglio ─────────────────────────────────────────────────────
def indice_stampa(wb):
    """{riga di SERRAMENTI: (limite DEI, codice, nome, blocco)} dal foglio Stampa."""
    ws = wb["Stampa"]
    fuori = {}
    for riga in ws.iter_rows(min_row=1, max_row=4000, max_col=8, values_only=True):
        if not riga or not testo(riga[0]).isdigit():
            continue
        r = num(riga[1])
        d = num(riga[3])
        if r is None or d is None:
            continue
        fuori[int(r)] = {"dei": d, "codice": testo(riga[7]), "nome": testo(riga[6]),
                         "blocco": testo(riga[2])[:1]}
    return fuori


def leggi_righe(wb, dia):
    """Le righe valorizzate dei blocchi A..D, in ordine di riga del foglio."""
    ws = wb["SERRAMENTI"]
    griglia = [r for r in ws.iter_rows(min_row=1, max_row=60, max_col=270, values_only=True)]
    intestazioni = griglia[4]  # riga 5: nome dell'accessorio per colonna
    stampa = indice_stampa(wb)

    for blocco, colonna in QUANTITA_LEGNO.items():
        for r in range(7, 57):
            if (num(cella(griglia, r, colonna)) or 0) > 0:
                sys.exit(f"blocco {blocco} (legno) valorizzato alla riga {r}: la mappa delle sue "
                         f"colonne non è verificata su un foglio reale, il caso sarebbe inventato")

    righe = []
    for blocco, cfg in BLOCCHI.items():
        # Prezzo fatturato del blocco: CHECK1 G6/G7/G8 somma tutta la colonna
        # (SERRAMENTI!I59 e sorelle) e l'operatore scrive spesso un importo
        # unico sulla prima riga, anche vuota. Va sulla prima riga raccolta:
        # al motore serve solo la somma (H39, «altri servizi» = 2 %).
        prezzo_blocco = sum(num(cella(griglia, r, cfg["prezzo"])) or 0 for r in range(7, 57))
        primo_del_blocco = True
        for r in range(7, 57):
            q = num(cella(griglia, r, cfg["quantita"]))
            if not q or q <= 0:
                continue
            etichetta_gruppo = testo(cella(griglia, r, cfg["gruppo"]))
            larghezza = num(cella(griglia, r, cfg["l"]))
            altezza = num(cella(griglia, r, cfg["h"]))
            dal_foglio = stampa.get(r, {})

            # Oscurante: nei blocchi B/F lo dichiara la colonna «Prodotto»
            # (SerrPers…), nel blocco C la riga È l'oscurante.
            oscurante_tipo, oscurante = None, None
            if blocco == "C":
                oscurante_tipo = OSCURANTE_SOLO.get(chiave(etichetta_gruppo))
                if not oscurante_tipo:
                    dia.avvisa(f"riga {r} blocco C: tipologia «{etichetta_gruppo}» sconosciuta, riga saltata")
                    continue
            else:
                oscurante_tipo = OSCURANTE_DEL_GRUPPO.get(chiave(etichetta_gruppo))
            gruppo_osc = {"tapparella": "avvolgibile", "persiana": "persiana", "scuro": "scuro"}
            if oscurante_tipo and oscurante_tipo in cfg["oscuranti"]:
                col, acc1, acc2 = cfg["oscuranti"][oscurante_tipo]
                nome_osc = testo(cella(griglia, r, col))
                if nome_osc:
                    candidati = [p for p in SEED["prodotti"] if p["gruppo"] == gruppo_osc[oscurante_tipo]]
                    scelto = scegli_per_nome(nome_osc, candidati, dia, "oscurante")
                    if not scelto:
                        dia.nome_non_mappato("prodotto oscurante", nome_osc, f"riga {r} blocco {blocco}")
                    oscurante = dict(prodotto=scelto, accessori=accessori_marcati(
                        griglia, intestazioni, r, acc1, acc2, gruppo_osc[oscurante_tipo], dia,
                        f"oscurante riga {r}"))
                elif blocco != "C":
                    dia.avvisa(f"riga {r} blocco {blocco}: il gruppo dichiara «{etichetta_gruppo}» ma "
                               f"nessun prodotto oscurante è scelto — il motore non sa mettere la riga "
                               f"nel gruppo B senza voce DEI dell'oscurante")

            # Prodotto della riga: il codice già risolto dal foglio «Stampa»
            # (classe del cassonetto, intervallo della schermatura) vince sul
            # nome del menu a tendina.
            nome_prodotto = testo(cella(griglia, r, cfg["tipologia"])) if cfg["tipologia"] else ""
            prodotto = None
            codice_stampa = dal_foglio.get("codice", "")
            if CODICE_DEI.match(codice_stampa):
                prodotto = prodotto_per_codice(codice_stampa)
                if not prodotto:
                    dia.avvisa(f"riga {r}: il codice «{codice_stampa}» di «Stampa» non è nel seed")
            if prodotto is None and blocco == "C" and oscurante and oscurante["prodotto"]:
                prodotto = oscurante["prodotto"]  # la riga è l'oscurante stesso
            if prodotto is None and nome_prodotto:
                prodotto = scegli_per_nome(nome_prodotto, SEED["prodotti"], dia, "prodotto")
            if prodotto is None:
                dia.nome_non_mappato("prodotto", nome_prodotto or codice_stampa or "(vuoto)",
                                     f"riga {r} blocco {blocco}")
                continue

            accessori = []
            if cfg["accessori"]:
                accessori = accessori_marcati(griglia, intestazioni, r, cfg["accessori"][0],
                                              cfg["accessori"][1], prodotto["gruppo"], dia,
                                              f"riga {r}")
            if oscurante:
                # Nel blocco C la riga È l'oscurante: i suoi accessori sono
                # quelli della riga, non di un oscurante «abbinato».
                accessori += oscurante["accessori"]
            if blocco == "C":
                oscurante_tipo, oscurante = None, None  # gli oscuranti soli non abbinano nulla

            categoria = categoria_di(prodotto)
            if categoria is None:
                dia.nome_non_mappato("categoria", prodotto["codice"],
                                     f"gruppo {prodotto['gruppo']}/{prodotto.get('famiglia')}")
                continue
            descrizione = prodotto["nome"]
            if oscurante_tipo:
                descrizione = f"{descrizione} + {oscurante_tipo}"
            if blocco == "B" and prodotto["gruppo"] == "cassonetto" and not oscurante_tipo:
                # Il foglio conta questo cassonetto nel massimale B (T9/Z9). Il
                # motore riconosce il blocco dall'oscurante dichiarato sulla
                # riga: è la tapparella che il cassonetto ospita, già prezzata
                # sulla riga del serramento e quindi senza voce DEI propria
                # (`oscuranteTipologia` resta vuota). La descrizione resta
                # quella del cassonetto.
                oscurante_tipo = "tapparella"
            righe.append({
                "_riga": r, "_blocco": blocco, "_deiFoglio": dal_foglio.get("dei"),
                "categoria": categoria,
                "tipologia": prodotto["codice"],
                "descrizione": descrizione,
                "quantita": int(q) if q == int(q) else q,
                "larghezzaMm": int(larghezza) if larghezza else None,
                "altezzaMm": int(altezza) if altezza else None,
                "prezzoTotCent": round(prezzo_blocco * 100) if primo_del_blocco and prezzo_blocco else None,
                "oscuranteIntegrato": oscurante_tipo,
                "oscuranteTipologia": (oscurante["prodotto"]["codice"]
                                       if oscurante and oscurante["prodotto"] else None),
                "accessori": accessori,
            })
            primo_del_blocco = False
    righe.sort(key=lambda x: (x["_riga"], x["_blocco"]))
    visti = {}
    for riga in righe:
        if riga["_riga"] in visti:
            dia.avvisa(f"riga {riga['_riga']}: valorizzata in due blocchi "
                       f"({visti[riga['_riga']]} e {riga['_blocco']}) — «Stampa» ne mostra una sola")
        visti[riga["_riga"]] = riga["_blocco"]
    return righe


def leggi_controtelai(check1, dia):
    """Righe di controtelaio da CHECK1 (D = misura dichiarata, E = prezzo)."""
    fuori = []
    for r in RIGHE_CONTROTELAIO:
        misura = num(cella(check1, r, "D")) or 0
        if misura <= 0:
            continue
        famiglia = testo(cella(check1, r, "B"))
        variante = testo(cella(check1, r, "C"))
        prezzo = num(cella(check1, r, "E"))
        voce = next((v for v in SEED["controtelai"]
                     if chiave(v["famiglia"]) == chiave(famiglia) and chiave(v["variante"]) == chiave(variante)), None)
        if voce is None:
            voce = next((v for v in SEED["controtelai"]
                         if prezzo is not None and abs(v["prezzo"] - prezzo) < 0.005), None)
        if voce is None:
            dia.nome_non_mappato("controtelaio", f"{famiglia} — {variante}", f"CHECK1 riga {r}")
            continue
        fuori.append({"categoria": "controtelaio", "tipologia": voce["codice"],
                      "descrizione": f"{voce['famiglia']} — {voce['variante']}",
                      "quantita": int(misura) if voce["unita"] == "cad" else 1,
                      "larghezzaMm": None, "altezzaMm": None, "prezzoTotCent": None,
                      "misuraDei": misura, "oscuranteIntegrato": None,
                      "oscuranteTipologia": None, "accessori": []})
    return fuori


def opera(codice):
    return next(o for o in SEED["opere"] if o["codice"] == codice)


def costruisci(percorso, nome, detrazione, pct, salta, dia):
    wb = openpyxl.load_workbook(percorso, data_only=True, read_only=True)
    if "SERRAMENTI" not in wb.sheetnames or "CHECK1" not in wb.sheetnames:
        sys.exit("non è una copia del foglio «CALCOLO NUOVI LIMITI»")
    inizio = [r for r in wb["INIZIO"].iter_rows(min_row=1, max_row=25, max_col=12, values_only=True)]
    check1 = [r for r in wb["CHECK1"].iter_rows(min_row=1, max_row=50, max_col=22, values_only=True)]

    zona = testo(cella(inizio, 11, "H")).upper()[:1] or None
    km = num(cella(inizio, 17, "E"))
    piano = num(cella(inizio, 19, "E"))
    righe = leggi_righe(wb, dia) + leggi_controtelai(check1, dia)

    def h(r):
        return num(cella(check1, r, "H")) or 0.0

    def g(r):
        return num(cella(check1, r, "G")) or 0.0

    # Opzioni: il foglio dichiara «fatturato» ciò che ha un importo in G.
    fatturata = {c: g(RIGA_OPERA[c]) > 0 for c in RIGA_OPERA}
    opzioni = {
        "rilievo": "foro" if fatturata["rilievo_foro"] or not fatturata["rilievo_pezzo"] else "pezzo",
        "speseProfessionali": fatturata["spese_professionali"],
        "eventuali": [c for c in OPERE_EVENTUALI if fatturata[c]],
    }
    inclusa_motore = {}
    for c in RIGA_OPERA:
        if c in ("rilievo_foro", "rilievo_pezzo"):
            inclusa_motore[c] = (opzioni["rilievo"] == c.split("_")[1])
        elif c == "spese_professionali":
            inclusa_motore[c] = opzioni["speseProfessionali"]
        elif opera(c)["gruppo"] == "eventuali":
            inclusa_motore[c] = c in opzioni["eventuali"]
        else:
            inclusa_motore[c] = opera(c)["inclusaDefault"]

    # Totali: gli stessi numeri del foglio sommati con la regola del motore.
    massimali = h(6) + h(7) + h(8)
    controtelai = h(15) + h(19)
    check1_atteso = massimali + controtelai + sum(h(RIGA_OPERA[c]) for c in RIGA_OPERA if inclusa_motore[c])
    dei = num(cella(check1, 6, "T")) or 0.0
    check2_atteso = dei + controtelai + sum(
        h(RIGA_OPERA[c]) for c in RIGA_OPERA if inclusa_motore[c] and not opera(c)["esclusaDaCheck2"])

    solo_foglio = [c for c in RIGA_OPERA if fatturata[c] and not inclusa_motore[c] and h(RIGA_OPERA[c]) > 0]
    solo_motore = [c for c in RIGA_OPERA if inclusa_motore[c] and not fatturata[c] and h(RIGA_OPERA[c]) > 0]

    voci = {"massimale_A": h(6), "massimale_B": h(7), "massimale_C": h(8)}
    for c in RIGA_OPERA:
        voci[c] = h(RIGA_OPERA[c])
    for i, riga in enumerate([r for r in righe if r["categoria"] != "controtelaio"], start=1):
        if riga.get("_deiFoglio") is not None:
            voci[f"dei_riga_{i}"] = riga["_deiFoglio"]

    pattuito = (num(cella(check1, 46, "G")) or 0.0) - g(RIGA_OPERA["spese_professionali"])
    if pattuito <= 0:
        pattuito = sum((r["prezzoTotCent"] or 0) for r in righe) / 100
        dia.avvisa("CHECK1 G46 a zero: pattuito preso dalla somma dei prezzi di riga")

    caso = {
        "nome": nome,
        "_fonte": nome,
        "parametri": {
            "zona": zona, "piano": int(piano) if piano else None,
            "distanzaKm": km if km else None,
            "pattuitoCent": round(pattuito * 100), "pattuitoTipo": "imponibile",
            "detrazioneTipo": detrazione, "detrazionePct": pct, "opzioni": opzioni,
        },
        "righe": [{k: v for k, v in r.items() if not k.startswith("_")} for r in righe],
        "attesi": {
            "voci": voci, "deiProdotti": dei,
            "check1": check1_atteso, "check2": check2_atteso,
            "tolleranzaCent": 1, "tolleranzaTotaliCent": 5,
            "foglio": {
                "check1": num(cella(check1, 46, "H")), "check2": num(cella(check1, 46, "T")),
                "opereSoloNelFoglio": solo_foglio, "opereSoloNelMotore": solo_motore,
                # H46/T46 sommano le spese professionali FATTURATE (G35), il
                # motore il loro limite (H35): prima della fattura non c'è altro.
                "speseProfessionaliFatturate": g(RIGA_OPERA["spese_professionali"]),
            },
        },
    }
    if salta:
        caso["salta"] = salta
    wb.close()
    return caso


def main():
    p = argparse.ArgumentParser(description="foglio «CALCOLO NUOVI LIMITI» compilato → caso d'oro")
    p.add_argument("foglio")
    p.add_argument("--nome", required=True, help="nome del caso, es. fattura-62-2026")
    p.add_argument("--detrazione", default="ecobonus",
                   choices=["ecobonus", "ristrutturazione", "nessuna"])
    p.add_argument("--pct", type=float, default=50)
    p.add_argument("--salta", default=None, help="motivo per cui il test salta questo caso")
    p.add_argument("--zitto", action="store_true", help="nessuna diagnostica su stderr")
    a = p.parse_args()

    dia = Diagnostica(a.zitto)
    if not a.zitto:
        print(f"== {a.nome}", file=sys.stderr)
    caso = costruisci(a.foglio, a.nome, a.detrazione, a.pct, a.salta, dia)
    caso["_diagnostica"] = {"avvisi": dia.avvisi, "nomiNonMappati": dia.nomi_non_mappati}
    print(json.dumps(caso, ensure_ascii=False, indent=1))


if __name__ == "__main__":
    main()
